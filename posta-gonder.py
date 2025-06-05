import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
import pytz
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import schedule
import time
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import concurrent.futures
import logging

# Logging setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configuration
CONFIG = {
    'timezone': pytz.timezone('Europe/Istanbul'),
    'email': {
        'address': "alijak5818@gmail.com",
        'password': "xfbc fuvy fonx kbxi",
        'recipient': "halimali58@hotmail.com",
        'smtp_server': "smtp.gmail.com",
        'smtp_port': 587
    },
    'symbols': ['A1CAP.IS', 'A1YEN.IS', 'CIMSA.IS', 'SISE.IS'],
    'timeframes': {
        '1h': '60d', '2h': '3mo', '4h': '120d', '1d': '1y', '1wk': '3y', '1mo': '10y'
    },
    'timeframes_tr': {
        '1h': 'Saatlik', '2h': '2 Saatlik', '4h': '4 Saatlik', '1d': 'Günlük', '1wk': 'Haftalık', '1mo': 'Aylık'
    },
    'tab_colors': {
        'Saatlik': '87CEEB', '2 Saatlik': '98FB98', '4 Saatlik': 'FFFFE0',
        'Günlük': 'FF9800', 'Haftalık': 'E6E6FA', 'Aylık': 'FFB6C1'
    },
    'supertrend': {
        'atr_period': 10,
        'factor': 3.0,
        'atrline': 1.5
    },
    'signal': {
        'min_confirm_bars': 2,
        'max_confirm_bars': 5,
        'proximity_threshold': 0.02,
        'lookback_period': 75
    },
    'schedule_time': "19:00"
}

def compute_supertrend(df, atr_period, factor, atrline):
    """Calculate Supertrend indicator."""
    df = df.copy()
    df['TR'] = np.maximum.reduce([
        df['High'] - df['Low'],
        (df['High'] - df['Close'].shift()).abs(),
        (df['Low'] - df['Close'].shift()).abs()
    ])
    df['ATR'] = df['TR'].rolling(window=atr_period).mean()
    df['hl2'] = (df['High'] + df['Low']) / 2
    df['upperband'] = df['hl2'] + factor * df['ATR']
    df['lowerband'] = df['hl2'] - factor * df['ATR']

    direction = np.zeros(len(df), dtype=int)
    supertrend = np.zeros(len(df))
    close = df['Close'].values
    upperband = df['upperband'].values
    lowerband = df['lowerband'].values

    if len(df) > 0 and not np.isnan(lowerband[0]):
        supertrend[0] = lowerband[0]
        direction[0] = 1

    for i in range(1, len(df)):
        prev_supertrend = supertrend[i - 1]
        prev_direction = direction[i - 1]
        if close[i] > prev_supertrend:
            direction[i] = 1
        elif close[i] < prev_supertrend:
            direction[i] = -1
        else:
            direction[i] = prev_direction
        supertrend[i] = max(lowerband[i], prev_supertrend) if direction[i] == 1 else min(upperband[i], prev_supertrend)

    df['supertrend'] = supertrend
    df['direction'] = direction
    df['upatrline'] = df['supertrend'] + atrline * df['ATR']
    df['dnatrline'] = df['supertrend'] - atrline * df['ATR']
    return df

def fetch_data(symbol, timeframe, period):
    """Fetch stock data for a given symbol and timeframe."""
    try:
        if timeframe == '2h':
            df = yf.download(symbol, period=period, interval="60m", progress=False, auto_adjust=False, timeout=10)
            if df.empty:
                logging.warning(f"{symbol} için 1 saatlik veri boş.")
                return None
            df.index = pd.to_datetime(df.index, utc=True).tz_convert(CONFIG['timezone'])
            df = df.resample('2h').agg({
                'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'
            }).dropna()
        else:
            df = yf.download(symbol, period=period, interval=timeframe, progress=False, auto_adjust=False, timeout=10)
            df.index = pd.to_datetime(df.index, utc=True).tz_convert(CONFIG['timezone'])

        if df.empty or len(df) < 60:
            logging.warning(f"{symbol} için yeterli veri yok (uzunluk: {len(df)}).")
            return None
        df['Symbol'] = symbol.replace('.IS', '')
        return df
    except Exception as e:
        logging.error(f"{symbol} için veri alınırken hata: {e}")
        return None

def get_signals(df, min_confirm_bars, max_confirm_bars, proximity_threshold, lookback_period):
    """Generate buy/sell signals based on Supertrend."""
    df = compute_supertrend(df, **CONFIG['supertrend'])
    direction = df['direction'].values
    close = df['Close'].values
    supertrend = df['supertrend'].values
    high = df['High'].values
    low = df['Low'].values
    bar_index = np.arange(len(df))

    turn_green = np.insert((direction[1:] < direction[:-1]), 0, False)
    turn_red = np.insert((direction[1:] > direction[:-1]), 0, False)
    last_turn_green = bar_index[turn_green][-1] if np.any(turn_green) else np.nan
    last_turn_red = bar_index[turn_red][-1] if np.any(turn_red) else np.nan

    bars_since_green = len(df) - 1 - last_turn_green if not np.isnan(last_turn_green) else np.nan
    bars_since_red = len(df) - 1 - last_turn_red if not np.isnan(last_turn_red) else np.nan

    ll2 = np.min(low[int(last_turn_green):]) if not np.isnan(last_turn_green) else np.nan
    hh1 = np.max(high[int(last_turn_red):]) if not np.isnan(last_turn_red) else np.nan
    ll2_75 = np.min(low[-lookback_period:]) if len(low) >= lookback_period else np.nan
    hh1_75 = np.max(high[-lookback_period:]) if len(high) >= lookback_period else np.nan

    last_signal = None
    if not np.isnan(last_turn_green) and min_confirm_bars <= bars_since_green <= max_confirm_bars:
        confirm_al = all(close[-1 - i] > supertrend[-1 - i] for i in range(min_confirm_bars))
        if confirm_al and turn_green[int(last_turn_green)]:
            last_signal = "AL"
    if not np.isnan(last_turn_red) and min_confirm_bars <= bars_since_red <= max_confirm_bars:
        confirm_sat = all(close[-1 - i] < supertrend[-1 - i] for i in range(min_confirm_bars))
        if confirm_sat and turn_red[int(last_turn_red)]:
            last_signal = "SAT"

    curr_close = close[-1] if len(close) > 0 else np.nan
    symbol = df['Symbol'].iloc[0]
    buy_row = sell_row = buy_signal = sell_signal = alarm_color = None

    if last_signal == "AL" and not np.isnan(ll2):
        price_str = f"{ll2:.2f} ({ll2_75:.2f})".replace('.', ',') if not np.isnan(ll2_75) else f"{ll2:.2f}".replace('.', ',')
        buy_signal = f"{symbol} - AL => {price_str} - Son: {curr_close:.2f}".replace('.', ',')
        alarm_color = 'green' if curr_close <= ll2 * (1 + proximity_threshold) else None
        buy_row = [symbol, "AL", price_str, f"{curr_close:.2f}".replace('.', ','), alarm_color]

    if last_signal == "SAT" and not np.isnan(hh1):
        price_str = f"{hh1:.2f} ({hh1_75:.2f})".replace('.', ',') if not np.isnan(hh1_75) else f"{hh1:.2f}".replace('.', ',')
        sell_signal = f"{symbol} - SAT => {price_str} - Son: {curr_close:.2f}".replace('.', ',')
        alarm_color = 'red' if curr_close >= hh1 * (1 - proximity_threshold) else None
        sell_row = [symbol, "SAT", price_str, f"{curr_close:.2f}".replace('.', ','), alarm_color]

    return buy_signal, sell_signal, buy_row, sell_row

def send_email(excel_file_name):
    """Send email with Excel attachment."""
    try:
        msg = MIMEMultipart()
        msg['From'] = CONFIG['email']['address']
        msg['To'] = CONFIG['email']['recipient']
        msg['Subject'] = f"Dip Tepe Tarama Sonuçları - {datetime.now(CONFIG['timezone']).strftime('%d-%m-%Y %H:%M')}"

        body = "Merhaba,\n\nEkli dosyada dip ve tepe tarama sonuçları bulunmaktadır.\n\nİyi günler,\nOtomatik Tarama Sistemi"
        msg.attach(MIMEText(body, 'plain'))

        with open(excel_file_name, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename= {excel_file_name}')
        msg.attach(part)

        with smtplib.SMTP(CONFIG['email']['smtp_server'], CONFIG['email']['smtp_port']) as server:
            server.starttls()
            server.login(CONFIG['email']['address'], CONFIG['email']['password'])
            server.sendmail(CONFIG['email']['address'], CONFIG['email']['recipient'], msg.as_string())
        logging.info(f"E-posta gönderildi: {excel_file_name}")
    except Exception as e:
        logging.error(f"E-posta gönderilirken hataProperties: {e}")

def format_worksheet(worksheet, timeframes_tr, buy_rows, sell_rows, now):
    """Format Excel worksheet with signals."""
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    hyperlink_font = Font(color="0000FF", underline="single")

    columns_buy = ["Sembol", "Sinyal", "Fiyat (Dip)", "Son Fiyat"]
    columns_sell = ["Sembol", "Sinyal", "Fiyat (Tepe)", "Son Fiyat"]
    combined_rows = [
        [f"📈 AL Sinyali ({now})", "", "", "", "", f"📉 SAT Sinyali ({now})", "", "", "", ""],
        columns_buy + [""] + columns_sell + [""]
    ]
    max_rows = max(len(buy_rows), len(sell_rows))
    for i in range(max_rows):
        buy_row = buy_rows[i] if i < len(buy_rows) else ["", "", "", "", ""]
        sell_row = sell_rows[i] if i < len(sell_rows) else ["", "", "", "", ""]
        combined_rows.append(buy_row[:4] + [""] + sell_row[:4] + [""])

    for i, row in enumerate(combined_rows, 1):
        for j, value in enumerate(row, 1):
            cell = worksheet.cell(row=i, column=j)
            cell.value = value
            cell.alignment = center_alignment
            if i == 1:
                if j <= 4:
                    cell.fill = light_green_fill
                elif j >= 6 and j <= 9:
                    cell.fill = light_red_fill
                cell.font = bold_font
            elif i == 2:
                cell.font = bold_font
            else:
                if j in [1, 6] and value:
                    cell.hyperlink = f"https://tr.tradingview.com/chart/?symbol=BIST:{value}"
                    cell.font = hyperlink_font
                if j == 2 and value == "AL" or (j == 4 and buy_rows[i-3][4] == 'green'):
                    cell.fill = light_green_fill
                if j == 7 and value == "SAT" or (j == 9 and sell_rows[i-3][4] == 'red'):
                    cell.fill = light_red_fill

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    worksheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
    worksheet.merge_cells(start_row=1, start_column=11, end_row=1, end_column=14)
    worksheet.cell(row=1, column=11).fill = yellow_fill

    worksheet.cell(row=2, column=11).value = "Sembol"
    worksheet.cell(row=2, column=12).value = "Sinyal"
    worksheet.cell(row=2, column=13).value = '=IF(L3="SAT","Fiyat (TEPE)",IF(L3="AL","Fiyat (DIP)",IF(L3="","Fiyat","")))'
    worksheet.cell(row=2, column=14).value = "Son Fiyat"
    for col in [11, 12, 13, 14]:
        worksheet.cell(row=2, column=col).font = bold_font
        worksheet.cell(row=2, column=col).alignment = center_alignment

    worksheet.cell(row=3, column=11).value = '=IF(K1="","",HYPERLINK("https://tr.tradingview.com/chart/?symbol=BIST:"&K1,K1))'
    worksheet.cell(row=3, column=11).font = hyperlink_font
    worksheet.cell(row=3, column=12).value = '=IF(K1="","",IFERROR(INDEX(B:B,MATCH(K1,A:A,0)),IFERROR(INDEX(G:G,MATCH(K1,F:F,0)),"")))'
    worksheet.cell(row=3, column=13).value = '=IF(K1="","",IFERROR(INDEX(C:C,MATCH(K1,A:A,0)),IFERROR(INDEX(H:H,MATCH(K1,F:F,0)),"")))'
    worksheet.cell(row=3, column=14).value = '=IF(K1="","",IFERROR(INDEX(D:D,MATCH(K1,A:A,0)),IFERROR(INDEX(I:I,MATCH(K1,F:F,0)),"")))'

    green_rule = FormulaRule(formula=['L3="AL"'], fill=light_green_fill)
    red_rule = FormulaRule(formula=['L3="SAT"'], fill=light_red_fill)
    green_rule_n3 = FormulaRule(
        formula=['AND(L3="AL", K1<>"", N3<>"", VALUE(SUBSTITUTE(N3,",","."))<=VALUE(LEFT(SUBSTITUTE(M3," (",""),FIND(",",SUBSTITUTE(M3," (",""))-1))*1.02)'],
        fill=light_green_fill
    )
    red_rule_n3 = FormulaRule(
        formula=['AND(L3="SAT", K1<>"", N3<>"", VALUE(SUBSTITUTE(N3,",","."))>=VALUE(LEFT(SUBSTITUTE(M3," (",""),FIND(",",SUBSTITUTE(M3," (",""))-1))*0.98)'],
        fill=light_red_fill
    )
    worksheet.conditional_formatting.add('L3', green_rule)
    worksheet.conditional_formatting.add('L3', red_rule)
    worksheet.conditional_formatting.add('N3', green_rule_n3)
    worksheet.conditional_formatting.add('N3', red_rule_n3)

    for col_idx in range(1, 15):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = 14.5 if col_idx in [3, 8, 13] else 2 if col_idx in [5, 10] else 10

def run_analysis():
    """Run the analysis and generate Excel report."""
    now = datetime.now(CONFIG['timezone'])
    excel_file_name = f"Dip_Tepe_Tarama_Tum_Zamanlar_{now.strftime('%d-%m-%Y_%H.%M')}.xlsx"
    any_signals = False

    if now.weekday() >= 5:
        logging.warning(f"Hafta sonu ({now.strftime('%d-%m-%Y %H:%M')}). Borsalar kapalı olabilir.")

    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        for tf, period in CONFIG['timeframes'].items():
            logging.info(f"İşleniyor: {CONFIG['timeframes_tr'][tf]}")
            buy_rows = []
            sell_rows = []

            with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
                future_to_symbol = {executor.submit(fetch_data, sym, tf, period): sym for sym in CONFIG['symbols']}
                for future in concurrent.futures.as_completed(future_to_symbol):
                    sym = future_to_symbol[future]
                    try:
                        df = future.result()
                        if df is None:
                            continue
                        buy_signal, sell_signal, buy_row, sell_row = get_signals(df, **CONFIG['signal'])
                        if buy_signal:
                            logging.info(f"AL Sinyali: {buy_signal}")
                            buy_rows.append(buy_row)
                            any_signals = True
                        if sell_signal:
                            logging.info(f"SAT Sinyali: {sell_signal}")
                            sell_rows.append(sell_row)
                            any_signals = True
                    except Exception as e:
                        logging.error(f"{sym} işlenirken hata: {e}")

            if buy_rows or sell_rows:
                worksheet = writer.book.create_sheet(CONFIG['timeframes_tr'][tf])
                worksheet.sheet_properties.tabColor = CONFIG['tab_colors'].get(CONFIG['timeframes_tr'][tf], 'FFFFFF')
                format_worksheet(worksheet, CONFIG['timeframes_tr'][tf], buy_rows, sell_rows, now.strftime("%d-%m-%Y %H:%M"))
            else:
                logging.info(f"{CONFIG['timeframes_tr'][tf]} için sinyal bulunamadı.")

        if not any_signals:
            logging.warning("Hiçbir sinyal bulunamadı.")
            pd.DataFrame([["Hiçbir sinyal bulunamadı"]], columns=["Bilgi"]).to_excel(writer, sheet_name="Bilgi", index=False)

    if any_signals:
        send_email(excel_file_name)
    logging.info("Analiz tamamlandı.")

if __name__ == "__main__":
    schedule.every().monday.at(CONFIG['schedule_time']).do(run_analysis)
    schedule.every().tuesday.at(CONFIG['schedule_time']).do(run_analysis)
    schedule.every().wednesday.at(CONFIG['schedule_time']).do(run_analysis)
    schedule.every().thursday.at(CONFIG['schedule_time']).do(run_analysis)
    schedule.every().friday.at(CONFIG['schedule_time']).do(run_analysis)

    logging.info(f"Zamanlayıcı başlatıldı: Hafta içi {CONFIG['schedule_time']}")
    run_analysis()  # Run immediately for testing
    while True:
        schedule.run_pending()
        time.sleep(60)
